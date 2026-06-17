"""
actualizar.py — Lee el Excel de la quiniela y genera datos.json
Uso: python actualizar.py
Requiere: pip install openpyxl
"""

import json, os
from datetime import datetime
from openpyxl import load_workbook

EXCEL = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Quiniela_Mundial2026_v3.xlsx")
JSON  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "datos.json")
JUGADORES = ['Yosember','Josmary','Carlos','Osmar','Jonathan','Sol','Mayra','Peter','Carolina','Jose']

print(f"Leyendo {EXCEL}...")
wb = load_workbook(EXCEL, data_only=True)

# ── Puntajes y pronósticos por jugador ────────────────────────────────────────
pronosticos = {}  # {jugador: [{id, pred_l, pred_v, pts, resultado}]}

for jug in JUGADORES:
    ws = wb[jug]
    pts_total = ws["E112"].value
    partidos_jug = []

    for i in range(5, 77):  # filas 5-76 = partidos de grupos
        row = [ws.cell(i, c).value for c in range(1, 11)]
        id_ = row[0]
        if not (id_ and str(id_).isdigit()):
            continue
        pred_l   = row[4]
        pred_v   = row[6]
        pts      = row[8]
        resultado = str(row[9]).strip() if row[9] else "Pendiente"

        partidos_jug.append({
            "id":       int(id_),
            "pred_l":   int(pred_l) if pred_l is not None else None,
            "pred_v":   int(pred_v) if pred_v is not None else None,
            "pts":      int(pts) if pts is not None else None,
            "resultado": resultado,
        })

    pronosticos[jug] = {
        "pts_total": int(pts_total) if pts_total else 0,
        "partidos":  partidos_jug,
    }
    print(f"  {jug}: {pronosticos[jug]['pts_total']} pts")

# ── Ranking ───────────────────────────────────────────────────────────────────
ranking_lista = sorted(JUGADORES, key=lambda j: pronosticos[j]['pts_total'], reverse=True)
max_pts = pronosticos[ranking_lista[0]]['pts_total']

ranking = []
for i, name in enumerate(ranking_lista, 1):
    pts = pronosticos[name]['pts_total']
    # acumulado partido a partido (para gráfica)
    acumulado = []
    total = 0
    for p in pronosticos[name]['partidos']:
        if p['pts'] is not None:
            total += p['pts']
            acumulado.append({"id": p['id'], "pts": total})
    ranking.append({
        "pos":       i,
        "name":      name,
        "pts":       pts,
        "diff":      pts - max_pts,
        "acumulado": acumulado,
    })

# ── Partidos con marcador desde Resultados ────────────────────────────────────
ws_res = wb["Resultados"]
partidos = []
for row in ws_res.iter_rows(min_row=2, values_only=True):
    cols = list(row) + [None]*8
    id_, grupo, jornada, local, goles_l, goles_v, visitante = cols[:7]
    if (isinstance(id_, (int,float)) and 1 <= id_ <= 72
            and isinstance(goles_l,(int,float)) and isinstance(goles_v,(int,float))
            and local and visitante):
        partidos.append({
            "id":        int(id_),
            "grupo":     str(grupo).strip(),
            "jornada":   int(jornada) if jornada else 0,
            "local":     str(local).strip(),
            "goles_l":   int(goles_l),
            "goles_v":   int(goles_v),
            "visitante": str(visitante).strip(),
        })

ultimos5   = list(reversed(partidos))[:5]
pendientes = 72 - len(partidos)

# ── Guardar ───────────────────────────────────────────────────────────────────
data = {
    "ok":           True,
    "actualizado":  datetime.now().strftime("%d/%m/%Y %H:%M"),
    "ranking":      ranking,
    "pronosticos":  pronosticos,
    "partidos":     partidos,
    "ultimos5":     ultimos5,
    "stats": {
        "jugados":    len(partidos),
        "pendientes": pendientes,
    }
}

with open(JSON, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"\n✓ datos.json generado — {len(partidos)} jugados, líder: {ranking[0]['name']} ({ranking[0]['pts']} pts)")
